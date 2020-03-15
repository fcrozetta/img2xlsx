import os
from flask import Flask,request,send_file
import openpyxl as xl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from PIL import Image
import base64
import io
from pathlib import Path

app = Flask(__name__)

@app.route("/",methods=["GET","POST"])
def main(data=None):
    if request.method == "GET":
        return "To use this program, send a POST request with the image in base64"
    if request.method == "POST":
        output = Path('output.xlsx')
        if output.is_file():
            os.remove('output.xlsx')
        data = request.data
        imgbytes = base64.b64decode(data)
        buf = io.BytesIO(imgbytes)
        img = Image.open(buf)
        pix = img.load()
        width,height = img.size
        wb = xl.Workbook()
        ws = wb.active
        for row in range(height):
            for col in range(width):
                columLetter = get_column_letter(col +1)
                ws.row_dimensions[row +1].heigth = 2.5
                ws.column_dimensions[columLetter].width = 1
                rgb = pix[col,row]
                hexColor = '{:02x}{:02x}{:02x}'.format(rgb[0],rgb[1],rgb[2])
                cell = ws.cell(row=row +1,column=col +1,value='')
                cell.fill = PatternFill(start_color=hexColor,end_color=hexColor,fill_type='solid')
        ws.sheet_view.zoomScale = 20
        wb.save(output.absolute())
        return send_file('output.xlsx',as_attachment=True)

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port)