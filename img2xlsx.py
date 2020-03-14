import openpyxl as xl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from PIL import Image

img = Image.open('img.png')
pix = img.load()

width,height = img.size

out = xl.Workbook()
ws = out.active
for row in range(height):
    for  col in range(width):
        columnLetter = get_column_letter(col +1)
        ws.row_dimensions[row + 1].height = 2.5
        ws.column_dimensions[columnLetter].width = 1
        rgb = pix[col,row]
        hexColor = '{:02x}{:02x}{:02x}'.format(rgb[0],rgb[1],rgb[2])
        cell = ws.cell(row=row +1 ,column=col +1,value='')
        cell.fill = PatternFill(start_color=hexColor, end_color=hexColor, fill_type='solid')
ws.sheet_view_zoomScale = 20
out.save('output.xlsx')
